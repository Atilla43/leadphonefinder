"""Эндпоинты кампаний."""

import io
import json
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile

from api.schemas.campaign import (
    CampaignActionResponse,
    CampaignCreateResponse,
    CampaignDetail,
    CampaignListResponse,
    CampaignSummary,
    RecipientItem,
    RecipientsResponse,
)
from core.config import settings
from core.deps import get_data_reader, get_outreach_manager
from services.db_data_reader import DbDataReader

router = APIRouter(prefix="/api/campaigns", tags=["campaigns"])


def _campaign_id(c: dict) -> str:
    return c.get("campaign_id") or str(c.get("user_id", ""))


def _to_summary(c: dict) -> CampaignSummary:
    recipients = c.get("recipients", [])
    sent = c.get("sent_count", 0)
    warm = sum(
        1 for r in recipients
        if r.get("status") in ("warm", "warm_confirmed")
    )
    replied = sum(
        1 for r in recipients
        if r.get("status") in ("talking", "warm", "warm_confirmed", "referral")
    )
    response_rate = (replied / sent * 100) if sent else 0.0
    conversion_rate = (warm / sent * 100) if sent else 0.0

    return CampaignSummary(
        campaign_id=_campaign_id(c),
        name=c.get("name", ""),
        user_id=c.get("user_id", 0),
        status=c.get("status", "unknown"),
        recipients_total=len(recipients),
        sent_count=sent,
        warm_count=c.get("warm_count", 0),
        rejected_count=c.get("rejected_count", 0),
        not_found_count=c.get("not_found_count", 0),
        response_rate=round(response_rate, 2),
        conversion_rate=round(conversion_rate, 2),
        has_system_prompt=bool(c.get("system_prompt")),
        has_service_info=bool(c.get("service_info")),
    )


@router.get("", response_model=CampaignListResponse)
async def list_campaigns(
    status: str | None = Query(default=None),
    reader: DbDataReader = Depends(get_data_reader),
) -> CampaignListResponse:
    """Список всех кампаний."""
    campaigns = await reader.get_all_campaigns()
    if status:
        allowed = {s.strip() for s in status.split(",")}
        campaigns = [c for c in campaigns if c.get("status") in allowed]
    return CampaignListResponse(campaigns=[_to_summary(c) for c in campaigns])


@router.get("/{campaign_id}", response_model=CampaignDetail)
async def get_campaign(
    campaign_id: str,
    reader: DbDataReader = Depends(get_data_reader),
) -> CampaignDetail:
    """Детали кампании."""
    c = await reader.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")

    recipients = c.get("recipients", [])
    breakdown: dict[str, int] = {}
    for r in recipients:
        s = r.get("status", "pending")
        breakdown[s] = breakdown.get(s, 0) + 1

    return CampaignDetail(
        campaign_id=_campaign_id(c),
        name=c.get("name", ""),
        user_id=c.get("user_id", 0),
        status=c.get("status", "unknown"),
        offer=c.get("offer", ""),
        system_prompt=c.get("system_prompt", ""),
        service_info=c.get("service_info", ""),
        manager_ids=c.get("manager_ids", []),
        work_hour_start=c.get("work_hour_start"),
        work_hour_end=c.get("work_hour_end"),
        recipients_total=len(recipients),
        sent_count=c.get("sent_count", 0),
        warm_count=c.get("warm_count", 0),
        rejected_count=c.get("rejected_count", 0),
        not_found_count=c.get("not_found_count", 0),
        statuses_breakdown=breakdown,
    )


@router.get("/{campaign_id}/recipients", response_model=RecipientsResponse)
async def list_recipients(
    campaign_id: str,
    status: str | None = Query(default=None),
    search: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    reader: DbDataReader = Depends(get_data_reader),
) -> RecipientsResponse:
    """Список получателей кампании с фильтрами."""
    c = await reader.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")

    recipients = c.get("recipients", [])

    # Фильтрация по статусу
    if status:
        allowed = {s.strip() for s in status.split(",")}
        recipients = [r for r in recipients if r.get("status") in allowed]

    # Поиск по имени/компании
    if search:
        q = search.lower()
        recipients = [
            r for r in recipients
            if q in (r.get("company_name", "")).lower()
            or q in (r.get("contact_name", "") or "").lower()
        ]

    total = len(recipients)
    page = recipients[offset: offset + limit]

    items = [
        RecipientItem(
            phone=r.get("phone", ""),
            company_name=r.get("company_name", ""),
            contact_name=r.get("contact_name"),
            category=r.get("category"),
            status=r.get("status", "pending"),
            last_message_at=r.get("last_message_at"),
            messages_count=len(r.get("conversation_history", [])),
            ping_count=r.get("ping_count", 0),
            rating=r.get("rating"),
            address=r.get("address"),
            account_phone=r.get("account_phone"),
        )
        for r in page
    ]

    return RecipientsResponse(recipients=items, total=total, offset=offset, limit=limit)


@router.post("/create", response_model=CampaignCreateResponse, status_code=201)
async def create_campaign(
    file: UploadFile = File(...),
    offer: str = Form(...),
    name: str = Form(""),
    service_info: str = Form(""),
    system_prompt: str = Form(""),
    user_id: int = Form(0),
    manager_ids: str = Form(""),
    work_hour_start: int | None = Form(None),
    work_hour_end: int | None = Form(None),
    reader: DbDataReader = Depends(get_data_reader),
) -> CampaignCreateResponse:
    """Создать кампанию из загруженного файла (CSV/XLSX)."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    # Если user_id не указан, используем owner из конфига
    if user_id == 0 and settings.owner_telegram_id:
        user_id = settings.owner_telegram_id

    # Парсим manager_ids из строки "123,456,789"
    parsed_manager_ids: list[int] = []
    if manager_ids.strip():
        for mid in manager_ids.split(","):
            mid = mid.strip()
            if mid.isdigit():
                parsed_manager_ids.append(int(mid))

    content = await file.read()
    items: list[dict] = []

    if file.filename.endswith(".csv"):
        items = _parse_csv(content)
    elif file.filename.endswith((".xlsx", ".xls")):
        items = _parse_excel(content)
    else:
        raise HTTPException(
            status_code=400,
            detail="Неподдерживаемый формат. Используйте .csv или .xlsx",
        )

    if not items:
        raise HTTPException(status_code=400, detail="Файл не содержит валидных данных")

    # Конвертируем в recipients
    recipients = [_to_recipient(item) for item in items]

    # Генерируем campaign
    campaign_id = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    campaign_name = name or offer[:40].replace("\n", " ")

    campaign_data = {
        "user_id": user_id,
        "offer": offer,
        "recipients": recipients,
        "campaign_id": campaign_id,
        "name": campaign_name,
        "status": "pending",
        "sent_count": 0,
        "warm_count": 0,
        "rejected_count": 0,
        "not_found_count": 0,
        "manager_ids": parsed_manager_ids,
        "system_prompt": system_prompt,
        "service_info": service_info,
    }

    # Сохраняем в БД через DbOutreachStorage
    from services.db_storage import DbOutreachStorage
    storage = DbOutreachStorage()

    # Конвертируем в OutreachCampaign для сохранения
    from services.outreach_manager import _ensure_bot_importable
    _ensure_bot_importable()
    from bot.models.outreach import OutreachCampaign, OutreachRecipient

    campaign_recipients = [
        OutreachRecipient.from_dict(r) for r in recipients
    ]
    campaign_obj = OutreachCampaign(
        user_id=user_id,
        offer=offer,
        recipients=campaign_recipients,
        campaign_id=campaign_id,
        name=campaign_name,
        status="pending",
        manager_ids=parsed_manager_ids,
        system_prompt=system_prompt,
        service_info=service_info,
    )
    campaign_obj.work_hour_start = work_hour_start
    campaign_obj.work_hour_end = work_hour_end
    storage.save(campaign_obj)

    return CampaignCreateResponse(
        campaign_id=campaign_id,
        recipients_count=len(recipients),
    )


# ── Парсинг файлов ──

_HEADER_MAP: dict[str, str] = {
    "телефон": "phone",
    "phone": "phone",
    "компания": "company_name",
    "название": "company_name",
    "company": "company_name",
    "company_name": "company_name",
    "контакт": "contact_name",
    "contact": "contact_name",
    "contact_name": "contact_name",
    "имя": "contact_name",
    "категория": "category",
    "category": "category",
    "рейтинг": "rating",
    "rating": "rating",
    "отзывы": "reviews_count",
    "reviews": "reviews_count",
    "сайт": "website",
    "website": "website",
    "адрес": "address",
    "address": "address",
    "директор": "director_name",
    "director": "director_name",
    "часы работы": "working_hours",
    "working_hours": "working_hours",
}


def _map_headers(raw_headers: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for idx, header in enumerate(raw_headers):
        key = header.strip().lower()
        if key in _HEADER_MAP:
            mapping[idx] = _HEADER_MAP[key]
    return mapping


def _normalize_phone(raw: str) -> str:
    digits = "".join(c for c in raw if c.isdigit())
    if len(digits) == 11 and digits[0] in ("7", "8"):
        digits = "7" + digits[1:]
    elif len(digits) == 10:
        digits = "7" + digits
    return "+" + digits if digits else raw


def _to_recipient(item: dict) -> dict:
    phone = _normalize_phone(item.get("phone", ""))
    return {
        "phone": phone,
        "company_name": item.get("company_name", ""),
        "contact_name": item.get("contact_name"),
        "category": item.get("category"),
        "rating": item.get("rating"),
        "reviews_count": item.get("reviews_count"),
        "website": item.get("website"),
        "address": item.get("address"),
        "director_name": item.get("director_name"),
        "status": "pending",
        "conversation_history": [],
        "ping_count": 0,
    }


def _parse_csv(content: bytes) -> list[dict]:
    import csv

    for encoding in ("utf-8", "cp1251", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        return []

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) < 2:
        return []

    col_map = _map_headers(rows[0])
    if "phone" not in col_map.values() or "company_name" not in col_map.values():
        if len(rows[0]) >= 2:
            col_map = {0: "phone", 1: "company_name"}

    items: list[dict] = []
    for row in rows[1:]:
        item: dict = {}
        for idx, field_name in col_map.items():
            if idx < len(row) and row[idx].strip():
                val = row[idx].strip()
                if field_name == "rating":
                    try:
                        item[field_name] = float(val.replace(",", "."))
                    except ValueError:
                        pass
                elif field_name == "reviews_count":
                    try:
                        item[field_name] = int(val)
                    except ValueError:
                        pass
                else:
                    item[field_name] = val
        if item.get("phone") and item.get("company_name"):
            items.append(item)

    return items


def _parse_excel(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h or "").strip() for h in rows[0]]
    col_map = _map_headers(headers)
    if "phone" not in col_map.values() or "company_name" not in col_map.values():
        if len(headers) >= 2:
            col_map = {0: "phone", 1: "company_name"}

    items: list[dict] = []
    for row in rows[1:]:
        item: dict = {}
        for idx, field_name in col_map.items():
            if idx < len(row) and row[idx] is not None:
                val = str(row[idx]).strip()
                if not val:
                    continue
                if field_name == "rating":
                    try:
                        item[field_name] = float(val.replace(",", "."))
                    except ValueError:
                        pass
                elif field_name == "reviews_count":
                    try:
                        item[field_name] = int(float(val))
                    except ValueError:
                        pass
                else:
                    item[field_name] = val
        if item.get("phone") and item.get("company_name"):
            items.append(item)

    wb.close()
    return items


# ─── Campaign Lifecycle ───


@router.post("/{campaign_id}/launch", response_model=CampaignActionResponse)
async def launch_campaign(
    campaign_id: str,
    mgr=Depends(get_outreach_manager),
) -> CampaignActionResponse:
    """Запускает кампанию (pending → sending)."""
    try:
        result = await mgr.launch_campaign(campaign_id)
        return CampaignActionResponse(**result)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=409, detail=str(e))


@router.post("/{campaign_id}/pause", response_model=CampaignActionResponse)
async def pause_campaign(
    campaign_id: str,
    mgr=Depends(get_outreach_manager),
) -> CampaignActionResponse:
    """Ставит кампанию на паузу (sending/listening → paused)."""
    try:
        result = await mgr.pause_campaign(campaign_id)
        return CampaignActionResponse(**result)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{campaign_id}/resume", response_model=CampaignActionResponse)
async def resume_campaign(
    campaign_id: str,
    mgr=Depends(get_outreach_manager),
) -> CampaignActionResponse:
    """Возобновляет кампанию (paused → sending/listening)."""
    try:
        result = await mgr.resume_campaign(campaign_id)
        return CampaignActionResponse(**result)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{campaign_id}/cancel", response_model=CampaignActionResponse)
async def cancel_campaign(
    campaign_id: str,
    mgr=Depends(get_outreach_manager),
) -> CampaignActionResponse:
    """Отменяет кампанию."""
    try:
        result = await mgr.cancel_campaign(campaign_id)
        return CampaignActionResponse(**result)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/{campaign_id}", response_model=CampaignActionResponse)
async def delete_campaign(
    campaign_id: str,
    reader: DbDataReader = Depends(get_data_reader),
) -> CampaignActionResponse:
    """Удаляет кампанию (только cancelled/completed)."""
    c = await reader.get_campaign(campaign_id)
    if not c:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if c.get("status") not in ("cancelled", "completed"):
        raise HTTPException(
            status_code=400,
            detail=f"Можно удалить только отменённые или завершённые кампании (текущий статус: {c.get('status')})",
        )

    from services.db_storage import DbOutreachStorage
    storage = DbOutreachStorage()
    storage.delete(c.get("user_id", 0), campaign_id)

    return CampaignActionResponse(
        campaign_id=campaign_id,
        status="deleted",
        message="Кампания удалена",
    )
